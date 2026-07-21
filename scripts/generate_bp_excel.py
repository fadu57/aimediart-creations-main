#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Génère le prévisionnel Excel AIMEDIArt — 3 scénarios × 36 mois (formules)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Paramètres communs (hypothèses BP) — aussi utilisés par generate_bp_charts.py ---
PRICES = {
    "ATELIER": 89.0,
    "HORIZON": 149.0,
    "ENVERGURE": 499.0,
    "RAYONNEMENT": 990.0,
}
RAYONNEMENT_ANNUAL_TTC = 10890.0  # 990 × 11
PAYING = ["ATELIER", "HORIZON", "ENVERGURE", "RAYONNEMENT"]

END_BASE = {
    1: {
        "ETINCELLE": 10,
        "ATELIER": 5,
        "HORIZON": 10,
        "ENVERGURE": 0,
        "RAYONNEMENT": 0,
        "ZENITH": 0,
    },
    2: {
        "ETINCELLE": 20,
        "ATELIER": 10,
        "HORIZON": 35,
        "ENVERGURE": 1,
        "RAYONNEMENT": 1,
        "ZENITH": 1,
    },
    3: {
        "ETINCELLE": 30,
        "ATELIER": 20,
        "HORIZON": 70,
        "ENVERGURE": 3,
        "RAYONNEMENT": 2,
        "ZENITH": 3,
    },
}

BILLING = {1: 0.60, 2: 0.50, 3: 0.40}
OVERAGE = {1: 0.05, 2: 0.10, 3: 0.12}
OPTS = {1: 20.0, 2: 25.0, 3: 30.0}

FIXED_START = 500.0
FIXED_Q_INC = 0.15
VAR_PCT = 0.01
CAPITAL = 2100.0
TVA = 0.20

CHURN_START = 9
CHURN_INITIAL = 0.05
CHURN_DECAY = 0.10
CATCHUP = 0.40

DEV_TJM = 450.0
DEV_DAYS_PER_MONTH = 10.0
DEV_MONTHLY_UNIT = DEV_TJM * DEV_DAYS_PER_MONTH

SCENARIOS = {
    "Prudent (-30%)": {"client_factor": 0.70, "dev_mode": "prudent"},
    "Base": {"client_factor": 1.00, "dev_mode": "base"},
    "Ambitieux (+30%)": {"client_factor": 1.30, "dev_mode": "ambitieux"},
}

ZENITH_SCHEDULE = {18: 15000.0, 30: 17000.0, 34: 17000.0}

# --- Références feuille Hypothèses (colonne B sauf tableaux) ---
P = "Hypothèses"
REF = {
    "capital": f"{P}!$B$2",
    "fixed_start": f"{P}!$B$3",
    "fixed_q": f"{P}!$B$4",
    "var_pct": f"{P}!$B$5",
    "tva": f"{P}!$B$6",
    "churn_start": f"{P}!$B$7",
    "churn_init": f"{P}!$B$8",
    "churn_decay": f"{P}!$B$9",
    "catchup": f"{P}!$B$10",
    "prix_atelier": f"{P}!$B$11",
    "prix_horizon": f"{P}!$B$12",
    "prix_envergure": f"{P}!$B$13",
    "prix_rayonnement": f"{P}!$B$14",
    "dev_tjm": f"{P}!$B$15",
    "dev_jours": f"{P}!$B$16",
    "dev_unit": f"{P}!$B$17",
    # Table facturation A1–A3 : lignes 20–22, col B = part mensuelle, C = overage, D = opts
    "bill_a1": f"{P}!$B$20",
    "bill_a2": f"{P}!$B$21",
    "bill_a3": f"{P}!$B$22",
    "ov_a1": f"{P}!$C$20",
    "ov_a2": f"{P}!$C$21",
    "ov_a3": f"{P}!$C$22",
    "opt_a1": f"{P}!$D$20",
    "opt_a2": f"{P}!$D$21",
    "opt_a3": f"{P}!$D$22",
    # Effectifs cibles fin d'année (base) : lignes 26–30, cols B/C/D = A1/A2/A3
    "cible_atelier": (26, "B"),
    "cible_horizon": (27, "B"),
    "cible_envergure": (28, "B"),
    "cible_rayonnement": (29, "B"),
    "cible_zenith_a2": f"{P}!$C$30",
    "cible_zenith_a3": f"{P}!$D$30",
    "zen_m18": f"{P}!$B$33",
    "zen_m30": f"{P}!$B$34",
}

FIRST_MONTH_ROW = 4
LAST_MONTH_ROW = 39
FACTOR_CELL = "$B$1"


def scaled_end(base: dict[int, dict[str, int]], factor: float) -> dict[int, dict[str, int]]:
    """Applique le facteur ±30 % — conservé pour generate_bp_charts.py."""
    out: dict[int, dict[str, int]] = {}
    for y, row in base.items():
        out[y] = {}
        for plan, n in row.items():
            out[y][plan] = max(0, round(n * factor))
    return out


def churn_rate(month: int) -> float:
    if month < CHURN_START:
        return 0.0
    return CHURN_INITIAL * ((1 - CHURN_DECAY) ** (month - CHURN_START))


def fixed_infra(month: int) -> float:
    q = (month - 1) // 3
    return FIXED_START * ((1 + FIXED_Q_INC) ** q)


def billing_factor(year: int) -> float:
    p = BILLING[year]
    return p + (1 - p) * (11 / 12)


def target_paying(end: dict, month: int) -> dict[str, float]:
    year = (month - 1) // 12 + 1
    m_in = (month - 1) % 12 + 1
    prev = end[year - 1] if year > 1 else {k: 0 for k in end[1]}
    curr = end[year]
    t = m_in / 12.0
    return {p: prev[p] + (curr[p] - prev[p]) * t for p in PAYING}


def dev_count_for_month(scenario_key: str, month: int) -> int:
    if scenario_key == "Prudent (-30%)":
        return 0
    if scenario_key == "Base":
        return 1 if month >= 13 else 0
    if scenario_key == "Ambitieux (+30%)":
        if month < 7:
            return 2
        if month < 13:
            return 3
        if month < 19:
            return 4
        return 5
    return 0


def run_scenario(name: str, end_targets: dict) -> list[dict]:
    """Simulation Python — utilisée par generate_bp_charts.py pour les graphiques."""
    state = {p: 0.0 for p in PAYING}
    cash = CAPITAL
    rows: list[dict] = []

    for month in range(1, 37):
        year = (month - 1) // 12 + 1
        churn = churn_rate(month)
        tgt = target_paying(end_targets, month)

        for p in PAYING:
            state[p] = state[p] * (1 - churn)
            if state[p] < tgt[p]:
                state[p] += (tgt[p] - state[p]) * CATCHUP
        if month % 12 == 0:
            for p in PAYING:
                state[p] = float(end_targets[year][p])

        paying = sum(state.values())
        sub_ttc = sum(state[p] * PRICES[p] for p in PAYING) * billing_factor(year)
        ov_ttc = paying * OVERAGE[year] * OPTS[year]
        zen_ttc = 0.0
        if month == 18 and end_targets[2]["ZENITH"] >= 1:
            zen_ttc = 15000.0
        elif month in (30, 34) and end_targets[3]["ZENITH"] >= 1:
            zen_ttc = 17000.0 * min(1.0, end_targets[3]["ZENITH"] / 3)

        rev_ttc = sub_ttc + ov_ttc + zen_ttc
        rev_ht = rev_ttc / (1 + TVA)
        fix = fixed_infra(month)
        dev_n = dev_count_for_month(name, month)
        dev_cost = dev_n * DEV_MONTHLY_UNIT
        var_cost = rev_ht * VAR_PCT
        ebitda = rev_ht - fix - dev_cost - var_cost
        cash += ebitda

        rows.append(
            {
                "Mois": month,
                "Année": year,
                "Atelier": round(state["ATELIER"], 1),
                "Horizon": round(state["HORIZON"], 1),
                "Envergure": round(state["ENVERGURE"], 1),
                "Rayonnement": round(state["RAYONNEMENT"], 1),
                "Clients payants": round(paying, 1),
                "MRR abo TTC": round(sub_ttc, 2),
                "Options TTC": round(ov_ttc, 2),
                "Zénith TTC": round(zen_ttc, 2),
                "CA TTC": round(rev_ttc, 2),
                "CA HT": round(rev_ht, 2),
                "Frais fixes infra": round(fix, 2),
                "Sous-traitance dev (HT)": round(dev_cost, 2),
                "Nb développeurs": dev_n,
                "Frais variables 1%": round(var_cost, 2),
                "EBITDA": round(ebitda, 2),
                "Trésorerie": round(cash, 2),
            }
        )
    return rows


def style_header(ws, row: int, ncol: int) -> None:
    fill = PatternFill("solid", fgColor="E63946")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _cible_interp(plan_row: int, year_cell: str, month_cell: str, factor: str) -> str:
    """Cible linéaire : prev + (curr-prev) × (mois dans l'année / 12), × facteur scénario."""
    b1, c1, d1 = f"{P}!$B${plan_row}", f"{P}!$C${plan_row}", f"{P}!$D${plan_row}"
    prev = f"IF({year_cell}=1,0,IF({year_cell}=2,{b1},IF({year_cell}=3,{c1},{d1})))"
    curr = f"IF({year_cell}=1,{b1},IF({year_cell}=2,{c1},IF({year_cell}=3,{d1},{d1})))"
    m_in_year = f"(({month_cell})-INT(({month_cell}-1)/12)*12)"
    return f"({prev}+({curr}-({prev}))*{m_in_year}/12)*{factor}"


def _stock_formula(plan_col: str, cible_col: str, row: int) -> str:
    """Parc clients : churn + rattrapage 40 % + snap fin d'année."""
    r = row
    prev = f"IF({r}={FIRST_MONTH_ROW},0,{plan_col}{r - 1})"
    a, i = f"$A{r}", f"$K{r}"
    churned = f"({prev})*(1-{i})"
    cible = f"{cible_col}{r}"
    catch = REF["catchup"]
    inner = (
        f"IF(MOD({a},12)=0,{cible},"
        f"{churned}+IF({churned}<{cible},({cible}-{churned})*{catch},0))"
    )
    return f"=ROUND({inner},1)"


def _dev_formula(dev_mode: str, row: int) -> str:
    a = f"$A{row}"
    if dev_mode == "prudent":
        return "=0"
    if dev_mode == "base":
        return f"=IF({a}>=13,1,0)"
    # ambitieux
    return f"=IF({a}<7,2,IF({a}<13,3,IF({a}<19,4,5)))"


def write_hypotheses_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Hypothèses", index=0)
    ws["A1"], ws["B1"] = "Paramètre", "Valeur"
    params = [
        ("Capital initial (€)", CAPITAL),
        ("Frais fixes infra départ (€/mois)", FIXED_START),
        ("Hausse frais fixes / trimestre", FIXED_Q_INC),
        ("Frais variables (% CA HT)", VAR_PCT),
        ("TVA (%)", TVA),
        ("Churn — début (mois)", CHURN_START),
        ("Churn initial (%/mois)", CHURN_INITIAL),
        ("Décroissance churn (%/mois)", CHURN_DECAY),
        ("Rattrapage rampe (%/mois)", CATCHUP),
        ("Prix Atelier TTC (€/mois)", PRICES["ATELIER"]),
        ("Prix Horizon TTC (€/mois)", PRICES["HORIZON"]),
        ("Prix L'Envergure TTC (€/mois)", PRICES["ENVERGURE"]),
        ("Prix Rayonnement TTC (€/mois)", PRICES["RAYONNEMENT"]),
        ("TJM développeur (€ HT)", DEV_TJM),
        ("Jours / mois / développeur", DEV_DAYS_PER_MONTH),
        ("Coût mensuel / développeur (€ HT)", f"=B15*B16"),
    ]
    for i, (label, val) in enumerate(params, start=2):
        ws.cell(i, 1, label)
        ws.cell(i, 2, val)

    ws["A19"] = "Facturation par année"
    ws["A20"], ws["B20"], ws["C20"], ws["D20"] = "Année", "Part mensuelle", "Part overage", "Panier opts (€)"
    for i, y in enumerate([1, 2, 3], start=20):
        ws.cell(i, 1, f"A{y}")
        ws.cell(i, 2, BILLING[y])
        ws.cell(i, 3, OVERAGE[y])
        ws.cell(i, 4, OPTS[y])

    ws["A24"] = "Effectifs cibles fin d'année (base, avant facteur scénario)"
    ws["A25"], ws["B25"], ws["C25"], ws["D25"] = "Plan", "Fin A1", "Fin A2", "Fin A3"
    plans = [
        ("Atelier", END_BASE[1]["ATELIER"], END_BASE[2]["ATELIER"], END_BASE[3]["ATELIER"]),
        ("Horizon", END_BASE[1]["HORIZON"], END_BASE[2]["HORIZON"], END_BASE[3]["HORIZON"]),
        ("L'Envergure", END_BASE[1]["ENVERGURE"], END_BASE[2]["ENVERGURE"], END_BASE[3]["ENVERGURE"]),
        ("Rayonnement", END_BASE[1]["RAYONNEMENT"], END_BASE[2]["RAYONNEMENT"], END_BASE[3]["RAYONNEMENT"]),
        ("Zénith (contrats)", 0, END_BASE[2]["ZENITH"], END_BASE[3]["ZENITH"]),
    ]
    for i, row in enumerate(plans, start=26):
        for j, val in enumerate(row):
            ws.cell(i, 1 + j, val)

    ws["A32"] = "Zénith — montants TTC (base)"
    ws["A33"], ws["B33"] = "Mois 18", ZENITH_SCHEDULE[18]
    ws["A34"], ws["B34"] = "Mois 30/34", ZENITH_SCHEDULE[30]

    ws["A36"] = "Note"
    ws["B36"] = "Modifier les valeurs ci-dessus : les 3 onglets scénario se recalculent automatiquement."

    style_header(ws, 1, 2)
    style_header(ws, 19, 4)
    style_header(ws, 24, 4)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14


def write_scenario_sheet(wb: Workbook, title: str, factor: float, dev_mode: str) -> str:
    """Crée un onglet scénario avec formules. Retourne le nom d'onglet (tronqué)."""
    sheet_name = title[:31]
    ws = wb.create_sheet(title=sheet_name)

    ws["A1"], ws["B1"] = "Facteur clients (scénario)", factor
    ws["A2"], ws["B2"] = "Zénith fin A2 (cible)", f"=ROUND({REF['cible_zenith_a2']}*{FACTOR_CELL},0)"
    ws["C2"] = "Zénith fin A3 (cible)"
    ws["D2"] = f"=ROUND({REF['cible_zenith_a3']}*{FACTOR_CELL},0)"

    headers = [
        "Mois",
        "Année",
        "Atelier",
        "Horizon",
        "Envergure",
        "Rayonnement",
        "Cible Atelier",
        "Cible Horizon",
        "Cible Enverg.",
        "Cible Rayonn.",
        "Taux churn",
        "Clients payants",
        "MRR abo TTC",
        "Options TTC",
        "Zénith TTC",
        "CA TTC",
        "CA HT",
        "Frais fixes infra",
        "Nb développeurs",
        "Sous-traitance dev (HT)",
        "Frais variables 1%",
        "EBITDA",
        "Trésorerie",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(FIRST_MONTH_ROW - 1, c, h)
    style_header(ws, FIRST_MONTH_ROW - 1, len(headers))

    f = FACTOR_CELL
    for row in range(FIRST_MONTH_ROW, LAST_MONTH_ROW + 1):
        a, b = f"A{row}", f"B{row}"
        ws[a] = row - FIRST_MONTH_ROW + 1
        ws[b] = f"=INT(({a}-1)/12)+1"

        # Cibles interpolées (colonnes G–J)
        ws[f"G{row}"] = f"={_cible_interp(26, b, a, f)}"
        ws[f"H{row}"] = f"={_cible_interp(27, b, a, f)}"
        ws[f"I{row}"] = f"={_cible_interp(28, b, a, f)}"
        ws[f"J{row}"] = f"={_cible_interp(29, b, a, f)}"

        # Taux churn
        ws[f"K{row}"] = (
            f"=IF({a}<{REF['churn_start']},0,"
            f"{REF['churn_init']}*(1-{REF['churn_decay']})^({a}-{REF['churn_start']}))"
        )

        # Stocks clients C–F
        ws[f"C{row}"] = _stock_formula("C", "G", row)
        ws[f"D{row}"] = _stock_formula("D", "H", row)
        ws[f"E{row}"] = _stock_formula("E", "I", row)
        ws[f"F{row}"] = _stock_formula("F", "J", row)

        # Clients payants
        ws[f"L{row}"] = f"=C{row}+D{row}+E{row}+F{row}"

        # Facteur facturation annuelle
        bill = (
            f"IF({b}=1,{REF['bill_a1']},IF({b}=2,{REF['bill_a2']},{REF['bill_a3']}))"
        )
        bill_f = f"({bill}+(1-({bill}))*11/12)"

        # MRR abonnements TTC
        ws[f"M{row}"] = (
            f"=(C{row}*{REF['prix_atelier']}+D{row}*{REF['prix_horizon']}"
            f"+E{row}*{REF['prix_envergure']}+F{row}*{REF['prix_rayonnement']})*{bill_f}"
        )

        # Options / dépassements
        ov = f"IF({b}=1,{REF['ov_a1']},IF({b}=2,{REF['ov_a2']},{REF['ov_a3']}))"
        opt = f"IF({b}=1,{REF['opt_a1']},IF({b}=2,{REF['opt_a2']},{REF['opt_a3']}))"
        ws[f"N{row}"] = f"=L{row}*{ov}*{opt}"

        # Zénith TTC
        ws[f"O{row}"] = (
            f"=IF({a}=18,IF($B$2>=1,{REF['zen_m18']},0),"
            f"IF(AND(OR({a}=30,{a}=34),$D$2>=1),"
            f"{REF['zen_m30']}*MIN(1,$D$2/3),0))"
        )

        ws[f"P{row}"] = f"=M{row}+N{row}+O{row}"
        ws[f"Q{row}"] = f"=P{row}/(1+{REF['tva']})"

        # Frais fixes infra (+15 % / trimestre)
        ws[f"R{row}"] = f"={REF['fixed_start']}*(1+{REF['fixed_q']})^INT(({a}-1)/3)"

        ws[f"S{row}"] = _dev_formula(dev_mode, row)
        ws[f"T{row}"] = f"=S{row}*{REF['dev_unit']}"
        ws[f"U{row}"] = f"=Q{row}*{REF['var_pct']}"
        ws[f"V{row}"] = f"=Q{row}-R{row}-T{row}-U{row}"

        if row == FIRST_MONTH_ROW:
            ws[f"W{row}"] = f"={REF['capital']}+V{row}"
        else:
            ws[f"W{row}"] = f"=W{row - 1}+V{row}"

    # Synthèse annuelle (formules SUMIF)
    syn = LAST_MONTH_ROW + 2
    ws.cell(syn, 1, "Synthèse annuelle").font = Font(bold=True, size=12)
    ws.cell(syn + 1, 1, "Année")
    for i, h in enumerate(["CA HT", "EBITDA", "Sous-traitance dev", "Trésorerie fin"], 2):
        ws.cell(syn + 1, i, h)

    br = f"$B${FIRST_MONTH_ROW}:$B${LAST_MONTH_ROW}"
    for y in [1, 2, 3]:
        r = syn + 1 + y
        ws.cell(r, 1, f"A{y}")
        ws.cell(r, 2, f"=SUMIF({br},{y},$Q${FIRST_MONTH_ROW}:$Q${LAST_MONTH_ROW})")
        ws.cell(r, 3, f"=SUMIF({br},{y},$V${FIRST_MONTH_ROW}:$V${LAST_MONTH_ROW})")
        ws.cell(r, 4, f"=SUMIF({br},{y},$T${FIRST_MONTH_ROW}:$T${LAST_MONTH_ROW})")
        last_m = y * 12
        last_row = FIRST_MONTH_ROW + last_m - 1
        ws.cell(r, 5, f"=W{last_row}")

    ws.cell(syn + 6, 1, "Effectifs fin A3 (cible × facteur)").font = Font(bold=True)
    ws.cell(
        syn + 7,
        1,
        f"=ROUND({P}!$D$26*{FACTOR_CELL},0)&\" Atelier · \"&ROUND({P}!$D$27*{FACTOR_CELL},0)"
        f"&\" Horizon · \"&ROUND({P}!$D$28*{FACTOR_CELL},0)&\" Enverg. · \"&ROUND({P}!$D$29*{FACTOR_CELL},0)&\" Rayonn.\"",
    )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    return sheet_name


def write_comparatif_sheet(wb: Workbook, sheet_names: dict[str, str]) -> None:
    """Comparatif annuel — formules vers les synthèses des onglets scénario."""
    ws = wb.create_sheet("Comparatif annuel")
    headers = [
        "Scénario",
        "CA HT A1",
        "EBITDA A1",
        "CA HT A2",
        "EBITDA A2",
        "CA HT A3",
        "EBITDA A3",
        "Trésorerie A3",
        "Dev A3/mois",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    syn_row = LAST_MONTH_ROW + 4  # 1re ligne données A1 dans la synthèse (après titre + en-têtes)

    for name, sn in sheet_names.items():
        q = f"'{sn}'" if " " in sn or "%" in sn else sn
        ws.append(
            [
                name,
                f"={q}!B{syn_row}",
                f"={q}!C{syn_row}",
                f"={q}!B{syn_row + 1}",
                f"={q}!C{syn_row + 1}",
                f"={q}!B{syn_row + 2}",
                f"={q}!C{syn_row + 2}",
                f"={q}!E{syn_row + 2}",
                f"={q}!S{LAST_MONTH_ROW}",
            ]
        )

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


def main() -> None:
    out = Path(__file__).resolve().parents[1] / "docs" / "business-plan-previsionnel-36m.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    write_hypotheses_sheet(wb)

    sheet_names: dict[str, str] = {}
    for name, cfg in SCENARIOS.items():
        sn = write_scenario_sheet(wb, name, cfg["client_factor"], cfg["dev_mode"])
        sheet_names[name] = sn

    write_comparatif_sheet(wb, sheet_names)

    try:
        wb.save(out)
    except PermissionError:
        out = out.with_name("business-plan-previsionnel-36m-new.xlsx")
        wb.save(out)

    print(f"Fichier généré (formules) : {out}")
    print("  → Modifier les hypothèses dans l'onglet « Hypothèses » pour recalculer les 3 scénarios.")


if __name__ == "__main__":
    main()
